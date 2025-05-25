import cv2
import numpy as np
import tensorflow as tf
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import img_to_array
import os
import pandas as pd
import datetime
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Get paths from environment variables or use defaults
root_dir = os.environ.get("IMS_INSTALLATION_DIR", os.path.dirname(os.path.abspath(__file__)))
models_dir = os.environ.get("IMS_MODELS_DIR", os.path.join(root_dir, "models"))

model_path = os.path.join(models_dir, "model.h5")
labels_path = os.path.join(models_dir, "labels1.txt")

logging.info(f"Root directory: {root_dir}")
logging.info(f"Models directory: {models_dir}")
logging.info(f"Model path: {model_path}")
logging.info(f"Labels path: {labels_path}")

# Create Excel root directory in the user-selected installation directory
excel_root = os.path.join(root_dir, "IMS EXCEL")
if not os.path.exists(excel_root):
    os.makedirs(excel_root)
    logging.info(f"Created Excel root directory: {excel_root}")

current_date = datetime.datetime.now()
year_month_folder = os.path.join(excel_root, f"{current_date.year}_{current_date.month:02d}")
if not os.path.exists(year_month_folder):
    os.makedirs(year_month_folder)
    logging.info(f"Created year-month folder: {year_month_folder}")

today_excel = os.path.join(year_month_folder, f"{current_date.day:02d}_{current_date.month:02d}_{current_date.year}.xlsx")
logging.info(f"Excel file path: {today_excel}")

detected_objects = []

registration_message = ""
registration_time = None
registration_display_duration = 1

# Load model with error handling
try:
    if not os.path.exists(model_path):
        logging.error(f"Model file not found: {model_path}")
        raise FileNotFoundError(f"Model file not found: {model_path}")
    
    model = load_model(model_path)
    logging.info(f"Model loaded successfully from: {model_path}")
except Exception as e:
    logging.error(f"Failed to load model: {e}")
    raise

# Load class names with error handling
class_names = {}
try:
    if not os.path.exists(labels_path):
        logging.error(f"Labels file not found: {labels_path}")
        raise FileNotFoundError(f"Labels file not found: {labels_path}")
    
    with open(labels_path, "r") as f:
        for line in f:
            if ": " in line:
                index, name = line.strip().split(": ", 1)
                class_names[int(index)] = name
    
    logging.info(f"Loaded {len(class_names)} class labels from: {labels_path}")
except Exception as e:
    logging.error(f"Failed to load class names: {e}")
    raise

def save_to_excel():
    if not detected_objects:
        logging.warning("No detected objects to save")
        return
    
    logging.info(f"Saving {len(detected_objects)} detected objects to Excel")
    
    try:
        df = pd.DataFrame(detected_objects, columns=["Timestamp", "Object", "Confidence"])
        logging.debug(f"Created DataFrame with {len(df)} rows")
        
        if os.path.exists(today_excel):
            logging.info(f"Excel file exists, appending data: {today_excel}")
            # Load existing data
            try:
                existing_wb = load_workbook(today_excel)
                
                # Check if Detections sheet exists
                if "Detections" in existing_wb.sheetnames:
                    existing_ws = existing_wb["Detections"]
                    existing_data = []
                    headers = [cell.value for cell in existing_ws[1]]
                    for row in existing_ws.iter_rows(min_row=2):
                        row_data = [cell.value for cell in row]
                        if any(cell is not None for cell in row_data):  # Skip empty rows
                            existing_data.append(row_data)
                    
                    if existing_data:
                        existing_df = pd.DataFrame(existing_data, columns=headers)
                        df = pd.concat([existing_df, df], ignore_index=True)
                        logging.info(f"Combined with existing data, total rows: {len(df)}")
                
                # Create new workbook with combined data
                wb = Workbook()
                ws1 = wb.active
                ws1.title = "Detections"
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws1.append(r)
                
                # Update summary with combined counts
                object_counts = Counter(df["Object"])
                summary_df = pd.DataFrame({
                    'Object': list(object_counts.keys()),
                    'Count': list(object_counts.values())
                })
                
                ws2 = wb.create_sheet(title="Summary")
                for r in dataframe_to_rows(summary_df, index=False, header=True):
                    ws2.append(r)
                
                # Save the updated workbook
                wb.save(today_excel)
                logging.info(f"Data appended to existing Excel file: {today_excel}")
                logging.info("Excel file updated successfully")
                print(f"Data appended to {today_excel}")
                
            except Exception as e:
                logging.error(f"Error appending to existing Excel file: {e}")
                # Create new file if there was an error with the existing one
                create_new_excel_file(df)
        else:
            logging.info(f"Creating new Excel file: {today_excel}")
            # Create new Excel file
            create_new_excel_file(df)
            
    except Exception as e:
        logging.error(f"Failed to save to Excel: {e}")
        print(f"Error saving to Excel: {e}")

def create_new_excel_file(df):
    try:
        object_counts = Counter(df["Object"])
        summary_df = pd.DataFrame({
            'Object': list(object_counts.keys()),
            'Count': list(object_counts.values())
        })
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Detections"
        for r in dataframe_to_rows(df, index=False, header=True):
            ws1.append(r)
            
        ws2 = wb.create_sheet(title="Summary")
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws2.append(r)
            
        wb.save(today_excel)
        logging.info(f"Data saved to new file {today_excel}")
        print(f"Data saved to new file {today_excel}")
        
    except Exception as e:
        logging.error(f"Failed to create new Excel file: {e}")
        print(f"Error creating Excel file: {e}")

# Initialize camera with error handling
try:
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        logging.error("Failed to open camera")
        raise RuntimeError("Failed to open camera")
    logging.info("Camera opened successfully")
except Exception as e:
    logging.error(f"Camera initialization failed: {e}")
    raise

cv2.namedWindow("IMS Feed", cv2.WINDOW_NORMAL)

logging.info("Starting Excel feed detection loop")

while True:
    ret, frame = cap.read()
    if not ret:
        logging.warning("Failed to read frame from camera")
        break

    image_resized = cv2.resize(frame, (224, 224))
    image_array = img_to_array(image_resized)
    image_array = np.expand_dims(image_array, axis=0)
    image_array /= 255.0

    try:
        prediction = model.predict(image_array, verbose=0)
        predicted_class = np.argmax(prediction)
        confidence = np.max(prediction) * 100

        class_label = class_names.get(predicted_class, "Unknown")
        
        color = (0, 255, 0) if confidence > 80 else (0, 165, 255)
        
        cv2.putText(frame, f"Class: {class_label}", (10, 30), 
                    cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2)
        cv2.putText(frame, f"Confidence: {confidence:.2f}%", (10, 70), 
                    cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2)
        
        current_time = datetime.datetime.now()
        if registration_message and registration_time:
            elapsed_time = (current_time - registration_time).total_seconds()
            if elapsed_time < registration_display_duration:
                text_size = cv2.getTextSize(registration_message, cv2.FONT_HERSHEY_SIMPLEX, 1, 2)[0]
                cv2.rectangle(frame, (frame.shape[1]//2 - text_size[0]//2 - 10, frame.shape[0]//2 - 20),
                             (frame.shape[1]//2 + text_size[0]//2 + 10, frame.shape[0]//2 + 20),
                             (0, 0, 0), -1)
                cv2.putText(frame, registration_message, 
                           (frame.shape[1]//2 - text_size[0]//2, frame.shape[0]//2 + 10),
                           cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            else:
                registration_message = ""
                registration_time = None
        
        cv2.putText(frame, "Press SPACE to register object", (10, frame.shape[0] - 20), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        cv2.putText(frame, "Press 'e' to exit", (10, frame.shape[0] - 50), 
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

        cv2.imshow("IMS Feed", frame)

        key = cv2.waitKey(1) & 0xFF
        
        if key == ord(' '):
            if confidence > 80 and class_label.lower() != 'noobject':
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                detected_objects.append([timestamp, class_label, confidence])
                logging.info(f"Object registered: {class_label} with confidence {confidence:.2f}% at {timestamp}")
                print(f"Registered: {class_label} with confidence {confidence:.2f}%")
                
                registration_message = f"REGISTERED: {class_label}"
                registration_time = datetime.datetime.now()
                
                save_to_excel()
            elif class_label.lower() == 'noobject':
                logging.info("'noobject' class ignored")
                print("'noobject' class ignored")
                registration_message = "'noobject' class ignored"
                registration_time = datetime.datetime.now()
            else:
                logging.info(f"Confidence too low ({confidence:.2f}%) to register")
                print(f"Confidence too low ({confidence:.2f}%) to register")
                registration_message = f"Confidence too low ({confidence:.2f}%)"
                registration_time = datetime.datetime.now()
        
        elif key == ord('e'):
            logging.info("Exit key pressed")
            break
            
    except Exception as e:
        logging.error(f"Error during prediction: {e}")
        continue

# Final save attempt
try:
    save_to_excel()
    logging.info("Final save completed")
except Exception as e:
    logging.error(f"Final save failed: {e}")

cap.release()
cv2.destroyAllWindows()
logging.info("Application closed")
